from PyQt5.QtWidgets import (
    QMainWindow, QSplitter, QListWidget, 
    QVBoxLayout, QWidget, QLabel, QPushButton,QDialog, QScrollArea,QListWidgetItem, QFileDialog  
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor

from raw_loader import load_sanitized_data, load_all_sanitized_sheets
from DealerInfoPanel import DealerInfoPanel
import pandas as pd
from NormalizerDialog import NormalizerDialog
import os,csv
from collections import defaultdict
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import Rule, CellIsRule
from openpyxl.utils import get_column_letter



class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Dealer-Personnel System")
        self.setGeometry(100, 100, 1000, 700)
        self.init_ui()
        self.load_data()
        self.load_mappings()  # Load mappings after data
        self.personnel_training_status = {}  


    
    def init_ui(self):
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout()
        main_widget.setLayout(main_layout)
        
        horizontal_splitter = QSplitter(Qt.Horizontal)
        
        left_sidebar = QSplitter(Qt.Vertical)
        
        self.raw_dealer_list = QListWidget()
        self.personnel_list = QListWidget()
        left_sidebar.addWidget(self.raw_dealer_list)
        left_sidebar.addWidget(self.personnel_list)
        


        self.dealer_info = DealerInfoPanel()

        self.personnel_name_label = QLabel("نام پرسنل:")
        self.personnel_position_label = QLabel("سمت:")
        # self.course_list_widget = QListWidget()


        # Replace personnel_details_label creation with:
        self.personnel_details_label = QLabel()
        self.personnel_details_label.setWordWrap(True)
        self.personnel_details_label.setTextFormat(Qt.RichText)

        self.dealer_details_label = QLabel()
        self.dealer_details_label.setWordWrap(True)
        self.dealer_details_label.setTextFormat(Qt.RichText)
        # Enable text selection
        self.personnel_details_label.setTextInteractionFlags(
            Qt.TextSelectableByMouse |  # Allow mouse selection
            Qt.LinksAccessibleByMouse   # Allow clicking links
        )
        
        # Improve selection visibility
        self.personnel_details_label.setStyleSheet(
            "QLabel::selected {"
            "   background-color: #3399ff;"
            "   color: white;"
            "}"
        )


        # self.personnel_details_label = QLabel()
        # self.personnel_details_label.setWordWrap(True)
        # self.personnel_details_label.setTextFormat(Qt.RichText)

        # self.personnel_details_label = QLabel()
        # self.personnel_details_label.setWordWrap(True)
        # self.personnel_details_label.setTextFormat(Qt.RichText)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(self.dealer_details_label)

        scroll_area2 = QScrollArea()
        scroll_area2.setWidgetResizable(True)
        scroll_area2.setWidget(self.personnel_details_label)



        self.categories_list = QListWidget()
        self.categories_list.setMaximumHeight(150)  # Limit height


        right_panels = QSplitter(Qt.Vertical)
        # right_panels.addWidget(self.personnel_name_label)
        # right_panels.addWidget(self.personnel_position_label)
        # right_panels.addWidget(self.categories_list)
        # right_panels.addWidget(QLabel("دوره‌ها:"))
        # right_panels.addWidget(self.course_list_widget)
        # Instead of course_list_widget, add personnel_details_label to right_panels
        # right_panels.addWidget(self.personnel_details_label)
        right_panels.addWidget(scroll_area)
        right_panels.addWidget(scroll_area2)

        horizontal_splitter.addWidget(left_sidebar)
        horizontal_splitter.addWidget(right_panels)
        horizontal_splitter.setSizes([200, 600])
        
        main_layout.addWidget(horizontal_splitter)
        
        self.raw_dealer_list.currentItemChanged.connect(self.show_dealer_info)
        self.personnel_list.currentItemChanged.connect(self.show_personnel_info)


        menubar = self.menuBar()
        settings_menu = menubar.addMenu('Settings')
        export_menu = menubar.addMenu('Export')
        mapping_action = settings_menu.addAction('Data Normalization')

        # Add export current dealer action
        export_current_action = export_menu.addAction('Export Current Dealer')
        export_current_action.triggered.connect(self.export_current_dealer)
        
        # Add export all dealers action
        export_all_action = export_menu.addAction('Export All Dealers')
        export_all_action.triggered.connect(self.export_all_dealers)
        
        mapping_action.triggered.connect(self.open_normalizer)

    def export_current_dealer(self):
        """Export detailed criteria data for the currently selected dealer"""
        current_item = self.raw_dealer_list.currentItem()
        if not current_item:
            return
            
        dealer_name = current_item.text()
        dealer_title = dealer_name[5:]
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Save Current Dealer Data",
            f"{dealer_title}_training_status.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
            
        self.export_dealer_data(dealer_name, filename)


    def export_dealer_data(self, dealer_name, filename):
        """Export dealer data to Excel file with formatting"""
        df = self.get_dealer_criteria_data(dealer_name)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Extract dealer title without code
            sheet_name = dealer_name.split(' - ')[-1][:30] if ' - ' in dealer_name else dealer_name[:30]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            format_worksheet(worksheet)

    def export_all_dealers(self):
        """Export detailed criteria data for all dealers with formatting"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Save All Dealers Data",
            "all_dealers_training_status.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
            
        all_dealers = sorted(self.raw['عنوان نمایندگی'].unique())
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for dealer_name in all_dealers:
                # Extract dealer title without code
                dealer_title = dealer_name.split(' - ')[-1] if ' - ' in dealer_name else dealer_name
                df = self.get_dealer_criteria_data(dealer_name)
                sheet_name = dealer_title[:25]  # Excel sheet name limit
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply formatting to each sheet
                worksheet = writer.sheets[sheet_name]
                format_worksheet(worksheet)


    def get_dealer_criteria_data(self, dealer_name):
        """Generate detailed criteria data for a dealer with proper logic for special cases"""
        categories = self.get_dealer_categories(dealer_name)
        dealer_personnel = self.raw[self.raw['عنوان نمایندگی'] == dealer_name]
        
        rows = []
        seen_combinations = set()
        
        for _, row in dealer_personnel.iterrows():
            name = row['نام و نام خانوادگی']
            main_pos = row.get('عنوان شغل', '')
            alt_pos = row.get('شغل موازی (ارتقا)', '')
            pcode = row.get('کد پرسنلی', '')
            rawcompany = row.get('company', '')
            
            positions_to_add = []
            if pd.notna(main_pos) and main_pos.strip():
                positions_to_add.append(main_pos.strip())
            if pd.notna(alt_pos) and alt_pos.strip():
                alt_positions = [p.strip() for p in alt_pos.split('&&&') if p.strip()]
                positions_to_add.extend(alt_positions)
            if not positions_to_add:
                positions_to_add = ['بدون سمت']
            
            for pos_text in positions_to_add:
                combination_key = (name, pos_text, pcode)
                if combination_key in seen_combinations:
                    continue
                seen_combinations.add(combination_key)
                
                # Skip unmapped positions (same as UI logic)
                if pos_text not in self.position_mapping:
                    continue
                
                mapped_company = self.company_mapping.get(rawcompany, rawcompany)
                mapped_position = self.position_mapping.get(pos_text, pos_text)
                mapped_categories = [self.car_mapping.get(cat, cat) for cat in categories]
                
                # Get passed courses
                filtered = self.raw[
                    (self.raw['کد پرسنلی'] == pcode) & 
                    (self.raw['عنوان نمایندگی'] == dealer_name)
                ]
                passed_courses = filtered['عنوان دوره'].dropna().unique().tolist()
                # Apply course mapping
                passed_courses = [self.course_mapping.get(c, c) for c in passed_courses]
                passed_set = set(passed_courses)
                
                # Get requirements
                after_requirements = self.get_matching_after_rows(
                    mapped_company, mapped_position, mapped_categories
                )
                sales_requirements = self.get_matching_sales_rows(
                    mapped_company, mapped_position
                )
                
                # Group requirements exactly like in show_personnel_info
                grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
                
                # Process after-sales requirements
                for req in after_requirements:
                    criteria = req.get("نام سرفصل", "").strip()
                    car = req.get("نام خودرو", "").strip() or "عمومی"
                    course = req.get("نام دوره آموزشی", "").strip()
                    
                    if criteria and criteria != "nan" and course and course != "nan":
                        grouped["after"][car][criteria].append(course)
                
                # Process sales requirements (from sales_sheets)
                sales_df = self.sales_sheets.get(mapped_company)
                if sales_df is not None:
                    for _, r in sales_df.iterrows():
                        row_pos = str(r.get("پست کاری", "")).strip()
                        if row_pos == mapped_position:
                            course = str(r.get("نام دوره آموزشی", "")).strip()
                            criteria = str(r.get("نام سرفصل", "")).strip()
                            car = "فروش"
                            
                            if criteria and criteria != "nan" and course and course != "nan":
                                grouped["sales"][car][criteria].append(course)
                
                # Apply the SAME logic as in show_personnel_info for determining pass status
                pass_status = defaultdict(lambda: defaultdict(dict))
                
                for file_name, cars in grouped.items():
                    for car, criteria_dict in cars.items():
                        # First pass: criteria passed if one course taken or name contains گازسوز
                        for crit, courses in criteria_dict.items():
                            passed = False
                            if any(c in passed_set for c in courses):
                                passed = True
                            elif "گازسوز" in crit:
                                passed = True
                            pass_status[file_name][car][crit] = passed

                        # Second pass: handle "ابزار مخصوص"
                        # If criteria contains "ابزار مخصوص", it passes only if ALL OTHER criteria in that car are passed
                        for crit in criteria_dict.keys():
                            if "ابزار مخصوص" in crit:
                                # Check if all other criteria in this car are passed
                                others = [c for c in criteria_dict.keys() if c != crit]
                                if all(pass_status[file_name][car].get(c, False) for c in others):
                                    pass_status[file_name][car][crit] = True
                                else:
                                    pass_status[file_name][car][crit] = False
                
                # Generate rows for export with proper logic applied
                for file_name, cars in grouped.items():
                    for car, criteria_dict in cars.items():
                        for crit, courses in criteria_dict.items():
                            # Skip empty criteria or courses (same as UI)
                            if not crit or not courses or all(not c for c in courses):
                                continue
                            
                            # Use the calculated pass_status
                            is_taken = pass_status[file_name][car].get(crit, False)
                            
                            # Determine the reason
                            if is_taken:
                                if "گازسوز" in crit:
                                    passed_course = "گازسوز (معاف)"
                                elif "ابزار مخصوص" in crit:
                                    passed_course = "ابزار مخصوص (شرطی)"
                                else:
                                    # Find the actual passed course
                                    passed_course = next((c for c in courses if c in passed_set), "تکمیل شده")
                            else:
                                if "ابزار مخصوص" in crit:
                                    passed_course = "ابزار مخصوص (سایر معیارها تکمیل نشده)"
                                else:
                                    passed_course = "گذرانده نشده"
                            
                            # Map file names to Persian
                            category_persian = "خدمات پس از فروش" if file_name == "after" else "فروش"
                            
                            rows.append({
                                'نمایندگی': dealer_name,
                                'نام پرسنل': name,
                                'سمت': pos_text,
                                'معیار': crit,
                                'دسته': category_persian,
                                'خودرو': car,
                                'گذرانده شده': 'بله' if is_taken else 'خیر',
                                'دلیل': passed_course
                            })
        
        return pd.DataFrame(rows)


    def open_normalizer(self):
        # Pass our data to the normalizer dialog
        dialog = NormalizerDialog(
            self, 
            self.raw, 
            self.dealers, 
            self.after_sheets, 
            self.sales_sheets
        )
        if dialog.exec_() == QDialog.Accepted:
            self.load_mappings()  # Reload mappings if saved
    
    def load_data(self):
        self.raw = load_sanitized_data("res/raw.xlsx")
        self.dealers = load_sanitized_data("res/dealers.xlsx")
        self.after =load_sanitized_data("res/after.xlsx")
        self.sales =load_sanitized_data("res/sales.xlsx")
        # Load all worksheets from after and sales files
        self.after_sheets = load_all_sanitized_sheets("res/after.xlsx")
        self.sales_sheets = load_all_sanitized_sheets("res/sales.xlsx")

        dealers = sorted(self.raw['عنوان نمایندگی'].unique())
        self.raw_dealer_list.addItems(dealers)


    def get_dealer_categories(self, dealer_name):
        """Extract categories for a specific dealer from dealers data"""
        categories = []
        
        # Find the dealer row in dealers dataframe
        dealer_row = self.dealers[self.dealers.iloc[:, 0] == dealer_name[:4]]  # Column A is dealer code
        if not dealer_row.empty:
            # Get the first matching row
            row = dealer_row.iloc[0]
            
            # Check columns D to AV (columns 3 to 47 in 0-based indexing)
            # Column D is index 3, Column AV is index 47
            for col_idx in range(3, min(48, len(row))):
                if col_idx < len(self.dealers.columns):
                    category_name = self.dealers.columns[col_idx]
                    cell_value = row.iloc[col_idx]
                    
                    # Check if cell contains 'p' (case insensitive)
                    if pd.notna(cell_value) and str(cell_value).strip().lower() == 'p':
                        categories.append(category_name)
        
        return categories

    def get_matching_sales_rows(self, company, position):
        """
        Find matching rows from 'sales_sheets' for given company and position.
        Sales sheets only use 'فروش' as car category by default.
        """
        matched_rows = []
        sheet_df = self.sales_sheets.get(company)
        if sheet_df is None:
            return []

        for _, row in sheet_df.iterrows():
            row_pos = str(row.get("پست کاری", "")).strip()
            if row_pos == position:
                matched_rows.append(row)

        print(f"[SALES MATCH] company={company}, position={position} → matched {len(matched_rows)} rows")
        return matched_rows


    def get_matching_after_rows(self, company, position, cars):
        """
        Find matching rows from 'after_sheets' for given company, position, and list of cars (categories).
        If any car is empty, treat it as 'عمومی'.
        """
        if not isinstance(cars, list):
            cars = [cars]

        # Clean car names, replace empty with "عمومی"
        cars = [str(c).strip() if str(c).strip() else "عمومی" for c in cars]
        cars.append("عمومی")  # Always allow default fallback
        cars = list(set(cars))  # Ensure uniqueness

        matched_rows = []

        sheet_df = self.after_sheets.get(company)
        if sheet_df is None:
            return []
        # if sheet_df is not None:
        #     print("Sheet columns:", sheet_df.columns.tolist())
        #     print(sheet_df[['نام خودرو', 'پست کاری']].dropna().head(10).to_string())


        for _, row in sheet_df.iterrows():
            row_car = str(row.get("نام خودرو", "")).strip() or "عمومی"
            row_pos = str(row.get("پست کاری", "")).strip()

            if row_car in cars and row_pos == position:
                matched_rows.append(row)

        # print(f"[AFTER MATCH] company={company}, position={position}, cars={cars} → matched {len(matched_rows)} rows")
        return matched_rows

    def load_mappings(self):
        self.position_mapping = {}
        self.car_mapping = {}
        self.company_mapping = {}
        
        # Load position mappings
        self.load_mapping_file('mappings/position_mapping.csv', self.position_mapping)
        # Load car category mappings
        self.load_mapping_file('mappings/car_mapping.csv', self.car_mapping)
        # Load company mappings
        self.load_mapping_file('mappings/company_mapping.csv', self.company_mapping)
        self.course_mapping = {}
        self.load_mapping_file('mappings/course_mapping.csv', self.course_mapping)

  
    def load_mapping_file(self, path, mapping_dict):
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader)  # Skip header
                for row in reader:
                    if len(row) >= 2:
                        mapping_dict[row[0]] = row[1]

    def generate_personnel_progress_summary(self):
        """
        Generate a summary table of each user's progress
        Returns: List of dictionaries with personnel progress data
        """
        summary_data = []
        after_totals = {'passed': 0, 'total': 0}
        sales_totals = {'passed': 0, 'total': 0}
        
        # Get all unique personnel from raw data
        personnel_records = {}
        
        for _, row in self.raw.iterrows():
            name = row['نام و نام خانوادگی']
            main_pos = row.get('عنوان شغل', '')
            alt_pos = row.get('شغل موازی (ارتقا)', '')
            pcode = row.get('کد پرسنلی', '')
            dealer_name = row['عنوان نمایندگی']
            
            # Handle position logic (same as in show_dealer_info)
            positions_to_add = []
            
            if pd.notna(main_pos) and main_pos.strip():
                positions_to_add.append(main_pos.strip())
            
            if pd.notna(alt_pos) and alt_pos.strip():
                alt_positions = [p.strip() for p in alt_pos.split('&&&') if p.strip()]
                positions_to_add.extend(alt_positions)
            
            if not positions_to_add:
                positions_to_add = ['بدون سمت']
            
            for pos_text in positions_to_add:
                combination_key = (name, pos_text, pcode)
                if combination_key not in personnel_records:
                    personnel_records[combination_key] = {
                        'name': name,
                        'position': pos_text,
                        'pcode': pcode,
                        'dealer_name': dealer_name
                    }
        
        # Process each personnel record
        for (name, position, pcode), record in personnel_records.items():
            mapped_position = self.position_mapping.get(position, position)
            
            # Skip unmapped positions (they appear as disabled in the UI)
            if position not in self.position_mapping:
                continue
                
            key = (pcode, mapped_position)
            
            # Get training status results for this personnel
            results = self.personnel_training_status.get(key, [])
            
            # Calculate progress for after and sales separately
            after_total = after_passed = 0
            sales_total = sales_passed = 0
            
            for r in results:
                if r['file'] == 'after':
                    after_total += 1
                    after_passed += 1 if r['is_taken'] else 0
                elif r['file'] == 'sales':
                    sales_total += 1
                    sales_passed += 1 if r['is_taken'] else 0
            
            # Calculate percentages
            after_percent = (after_passed / after_total * 100) if after_total > 0 else None
            sales_percent = (sales_passed / sales_total * 100) if sales_total > 0 else None
            
            # Format progress strings
            after_progress = f"{after_percent:.1f}% ({after_passed} of {after_total})" if after_percent is not None else "—"
            sales_progress = f"{sales_percent:.1f}% ({sales_passed} of {sales_total})" if sales_percent is not None else "—"
            
            # Add to totals for overall calculation
            after_totals['passed'] += after_passed
            after_totals['total'] += after_total
            sales_totals['passed'] += sales_passed
            sales_totals['total'] += sales_total
            
            summary_data.append({
                'name': name,
                'position': position,
                'after_progress': after_progress,
                'sales_progress': sales_progress,
                'dealer_name': record['dealer_name']
            })
        
        # Sort by name
        summary_data.sort(key=lambda x: x['name'])
        
        # Calculate total progress
        total_after_percent = (after_totals['passed'] / after_totals['total'] * 100) if after_totals['total'] > 0 else 0
        total_sales_percent = (sales_totals['passed'] / sales_totals['total'] * 100) if sales_totals['total'] > 0 else 0
        
        return summary_data, {
            'after_total': f"{total_after_percent:.1f}% ({after_totals['passed']} of {after_totals['total']})",
            'sales_total': f"{total_sales_percent:.1f}% ({sales_totals['passed']} of {sales_totals['total']})"
        }

    def print_personnel_progress_table(self):
        """
        Print a formatted table of personnel progress
        """
        summary_data, totals = self.generate_personnel_progress_summary()
        
        # Print table header
        print(f"{'User Name':<25} | {'Position':<20} | {'After Progress':<25} | {'Sales Progress':<25}")
        print("-" * 100)
        
        # Print each personnel row
        for record in summary_data:
            print(f"{record['name']:<25} | {record['position']:<20} | {record['after_progress']:<25} | {record['sales_progress']:<25}")
        
        # Print totals
        print("-" * 100)
        print(f"{'TOTAL':<25} | {'ALL POSITIONS':<20} | {totals['after_total']:<25} | {totals['sales_total']:<25}")

    def export_personnel_progress_csv(self, filename="personnel_progress_summary.csv"):
        """
        Export personnel progress summary to CSV file
        """
        import csv
        
        summary_data, totals = self.generate_personnel_progress_summary()
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['User Name', 'Position', 'Dealer', 'After Progress', 'Sales Progress']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            
            for record in summary_data:
                writer.writerow({
                    'User Name': record['name'],
                    'Position': record['position'],
                    'Dealer': record['dealer_name'],
                    'After Progress': record['after_progress'],
                    'Sales Progress': record['sales_progress']
                })
            
            # Add totals row
            writer.writerow({
                'User Name': 'TOTAL',
                'Position': 'ALL POSITIONS',
                'Dealer': 'ALL DEALERS',
                'After Progress': totals['after_total'],
                'Sales Progress': totals['sales_total']
            })
        
        print(f"Progress summary exported to {filename}")

    def show_personnel_info(self, item):
        if not item:
            return
        if not item.flags() & Qt.ItemIsSelectable:
        # Ignore disabled item
            return
        
        text = item.text()
        if "|" not in text:
            return
        
        dealer_code, name, position, pcode = [part.strip() for part in text.split("|")]
        mapped_position = self.position_mapping.get(position, position)
        key = (pcode, mapped_position)

        # Get training status results for this personnel
        results = self.personnel_training_status.get(key, [])

        
        # Get dealer full name
        dealer_full_name = next((d for d in self.raw['عنوان نمایندگی'].unique() 
                            if d.startswith(dealer_code)), "")
        


        # Find full dealer name and company
        dealer_full_name = next((d for d in self.raw['عنوان نمایندگی'].unique() if d.startswith(dealer_code)), "")
        filtered = self.raw[(self.raw['کد پرسنلی'] == pcode) & (self.raw['عنوان نمایندگی'] == dealer_full_name)]
        company = filtered.iloc[0].get('company', '') if not filtered.empty else ""

        passed_courses = filtered['عنوان دوره'].dropna().unique().tolist()
        passed_courses = [self.course_mapping.get(c, c) for c in passed_courses]
        passed_set = set(passed_courses)

        def colored_courses(courses, passed_set):
            parts = []
            for c in courses:
                # Skip empty or invalid courses
                if not c or str(c).strip() == "" or str(c).strip() == "nan":
                    continue
                if c in passed_set:
                    parts.append(f'<span style="background-color:#a2c1d5;">{c}</span>')  # light green
                else:
                    parts.append(f'<span style="background-color:#d3d3d3;">{c}</span>')  # grey
            return " &mdash; ".join(parts)

        passed_courses_str = colored_courses(passed_courses, passed_set)

        # Group results by file -> car -> criteria
        grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        for r in results:
            file_key = r.get("file", "after")
            grouped[file_key][r["car"]][r["criteria"]].append(r["course"])

        # Add sales courses for this person from sales_sheets
        mapped_company = self.company_mapping.get(company, company)
        sales_df = self.sales_sheets.get(mapped_company)
        if sales_df is not None:
            for _, row in sales_df.iterrows():
                row_pos = str(row.get("پست کاری", "")).strip()
                if row_pos == mapped_position:
                    # Swap the columns - A is actually course, C is actually criteria
                    course = str(row.get("نام دوره آموزشی", "")).strip()  # Column A
                    criteria = str(row.get("نام سرفصل", "")).strip()      # Column C
                    car = "فروش"
                    
                    # Skip empty or invalid criteria/courses
                    if criteria and criteria != "nan" and course and course != "nan":
                        grouped["sales"][car][criteria].append(course)

        # --- New: Determine criteria passed or not ---
        # Criteria is passed if:
        # 1) at least one course taken for it
        # 2) OR criteria includes "گازسوز"
        # 3) OR criteria includes "ابزار مخصوص" and all other criteria in the same car category are passed
        
        # We'll create a dict: pass_status[file][car][criteria] = True/False
        pass_status = defaultdict(lambda: defaultdict(dict))

        for file_name, cars in grouped.items():
            for car, criteria_dict in cars.items():
                # First pass: criteria passed if one course taken or name contains گازسوز
                for crit, courses in criteria_dict.items():
                    passed = False
                    if any(c in passed_set for c in courses):
                        passed = True
                    elif "گازسوز" in crit:
                        passed = True
                    pass_status[file_name][car][crit] = passed

                # Second pass: handle "ابزار مخصوص"
                # If criteria contains "ابزار مخصوص", it passes only if ALL OTHER criteria in that car are passed
                for crit in criteria_dict.keys():
                    if "ابزار مخصوص" in crit:
                        # Check if all other criteria in this car are passed
                        others = [c for c in criteria_dict.keys() if c != crit]
                        if all(pass_status[file_name][car].get(c, False) for c in others):
                            pass_status[file_name][car][crit] = True
                        else:
                            pass_status[file_name][car][crit] = False

        # --- Calculate pass counts and percentages per car and overall ---
        # count total and passed criteria per car and overall
        overall_total = 0
        overall_passed = 0

        criteria_html_parts = []
        for file_name, cars in grouped.items():
            file_label = "فروش" if file_name == "sales" else "خدمات پس از فروش"
            criteria_html_parts.append(f'<h1>{file_label}:</h1><br>')

            for car, crits in cars.items():
                total_criteria = len(crits)
                passed_criteria = sum(1 for crit in crits if pass_status[file_name][car].get(crit, False))

                overall_total += total_criteria
                overall_passed += passed_criteria

                percent = (passed_criteria / total_criteria * 100) if total_criteria else 0

                # Car header with passed/total and percent
                criteria_html_parts.append(
                    f'<div style="margin-right:20px;"><b>خودرو: {car} - '
                    f'{passed_criteria} از {total_criteria} (٪{percent:.1f})</b></div><br>'
                )
            
                for crit, courses in crits.items():
                    # Skip empty criteria or courses
                    if not crit or not courses or all(not c for c in courses):
                        continue
                        
                    passed = pass_status[file_name][car].get(crit, False)
                    bgcolor = "#a8d5a2" if passed else "#f8a5a5"  # light green or light red
                    course_html = colored_courses(courses, passed_set)
                    
                    # Skip if course_html is empty or only contains dashes/spaces
                    if not course_html.strip() or course_html.strip() == "&mdash;":
                        continue
                        
                    criteria_html_parts.append(
                        f'<div style="margin-right:40px; background-color:{bgcolor}; padding:4px; border-radius:4px; margin-bottom:3px;">'
                        f'{crit} ({course_html})</div>'
                    )
                criteria_html_parts.append('<br>')

        overall_percent = (overall_passed / overall_total * 100) if overall_total else 0

        # Generate progress summary table for the selected dealer
        

        # Add total pass info at top
        top_info = (
            f'<b>نام پرسنل:</b> {name} &mdash; '
            f'موارد گذرانده شده: {overall_passed} از {overall_total} (٪{overall_percent:.1f})<br>'
            f'<b>سمت:</b> {position}<br>'
            f'<b>نمایندگی:</b> {dealer_full_name}<br>'
            f'<b>شرکت:</b> {company}<br><br>'
        )

        criteria_section = "".join(criteria_html_parts) if criteria_html_parts else "— موردی یافت نشد —"

        html = (
            top_info +
            f'<b>دوره‌های الزامی و سرفصل‌ها:</b><br>{criteria_section}<br><br>' +
            f'<b>دوره‌های گذرانده شده:</b><br>{passed_courses_str}'
        )

        
        
        self.personnel_details_label.setText(html)


    # Add to MainWindow class in main_window.py
    def calculate_dealer_progress(self, dealer_name):
        """Calculate overall progress for a dealer's personnel"""
        dealer_personnel = self.raw[self.raw['عنوان نمایندگی'] == dealer_name]
        after_progress = []
        sales_progress = []
        
        for _, row in dealer_personnel.iterrows():
            pcode = row.get('کد پرسنلی', '')
            main_pos = row.get('عنوان شغل', '')
            mapped_position = self.position_mapping.get(main_pos, main_pos)
            
            if (pcode, mapped_position) in self.personnel_training_status:
                results = self.personnel_training_status[(pcode, mapped_position)]
                after_total = after_passed = 0
                sales_total = sales_passed = 0
                
                for r in results:
                    if r['file'] == 'after':
                        after_total += 1
                        after_passed += 1 if r['is_taken'] else 0
                    elif r['file'] == 'sales':
                        sales_total += 1
                        sales_passed += 1 if r['is_taken'] else 0
                
                if after_total > 0:
                    after_progress.append(after_passed / after_total)
                if sales_total > 0:
                    sales_progress.append(sales_passed / sales_total)
        
        # Calculate averages
        after_avg = sum(after_progress) / len(after_progress) * 100 if after_progress else 0
        sales_avg = sum(sales_progress) / len(sales_progress) * 100 if sales_progress else 0
        
        return after_avg, sales_avg


          
    def show_dealer_info(self, item):
        if item:
            dealer_name = item.text()
            self.dealer_info.display_info(dealer_name, self.raw)

            # Clear and populate categories list
            self.categories_list.clear()
            categories = self.get_dealer_categories(dealer_name)
            for category in categories:
                self.categories_list.addItem(category)

            dealer_personnel = self.raw[self.raw['عنوان نمایندگی'] == dealer_name]
            self.personnel_list.clear()

            personnel_entries = []
            seen_combinations = set()
            
            # Clear previous stored training status
            self.personnel_training_status = {}

            for _, row in dealer_personnel.iterrows():
                name = row['نام و نام خانوادگی']
                main_pos = row.get('عنوان شغل', '')
                alt_pos = row.get('شغل موازی (ارتقا)', '')
                pcode = row.get('کد پرسنلی', '')
                rawcompany = row.get('company')

                # Handle position logic
                positions_to_add = []

                if pd.notna(main_pos) and main_pos.strip():
                    positions_to_add.append(main_pos.strip())
                
                if pd.notna(alt_pos) and alt_pos.strip():
                    alt_positions = [p.strip() for p in alt_pos.split('&&&') if p.strip()]
                    positions_to_add.extend(alt_positions)
                
                if not positions_to_add:
                    positions_to_add = ['بدون سمت']

                for pos_text in positions_to_add:
                    combination_key = (name, pos_text, pcode)
                    if combination_key not in seen_combinations:
                        seen_combinations.add(combination_key)

                        mapped_company = self.company_mapping.get(rawcompany, rawcompany)
                        mapped_position = self.position_mapping.get(pos_text, pos_text)
                        mapped_categories = [self.car_mapping.get(cat, cat) for cat in categories]

                        # Get required trainings from after sheets
                        after_requirements = self.get_matching_after_rows(
                            mapped_company,
                            mapped_position,
                            mapped_categories
                        )

                        # Get required trainings from sales sheets
                        sales_requirements = self.get_matching_sales_rows(
                            mapped_company,
                            mapped_position
                        )

                        # Get passed courses for this personnel
                        passed_courses = self.raw[
                            (self.raw['کد پرسنلی'] == pcode)
                        ]['عنوان دوره'].dropna().unique().tolist()
                        taken_set = set(passed_courses)

                        # Compare and tag results
                        results = []

                        for req in after_requirements:
                            criteria = req.get("نام سرفصل", "").strip()
                            car = req.get("نام خودرو", "").strip() or "عمومی"
                            course = req.get("نام دوره آموزشی", "").strip()
                            is_taken = course in taken_set

                            results.append({
                                "file": "after",
                                "criteria": criteria,
                                "car": car,
                                "course": course,
                                "is_taken": is_taken
                            })

                        for req in sales_requirements:
                            criteria = req.get("معیار", "").strip()
                            car = "فروش"
                            course = req.get("نام دوره", "").strip()
                            is_taken = course in taken_set

                            results.append({
                                "file": "sales",
                                "criteria": criteria,
                                "car": car,
                                "course": course,
                                "is_taken": is_taken
                            })


                        self.personnel_training_status[(pcode, mapped_position)] = results

                        display_text = f"{dealer_name[:4]} | {name} | {pos_text} | {pcode}"

                        if pos_text not in self.position_mapping:
                            # Create QListWidgetItem disabled (not selectable)
                            item = QListWidgetItem(display_text)
                            item.setFlags(Qt.ItemIsEnabled)  # Enabled but NOT selectable (remove ItemIsSelectable)
                            item.setForeground(QColor('gray'))  # visually gray out
                            self.personnel_list.addItem(item)
                        else:
                            # Normal selectable item
                            item = QListWidgetItem(display_text)
                            self.personnel_list.addItem(item)
                        personnel_entries.append(display_text)

            # Sort personnel entries by name
            personnel_entries.sort(key=lambda x: x.split('|')[1].strip())

            for entry in personnel_entries:
                self.personnel_list.addItem(entry)

        summary_html = self.generate_dealer_progress_table_html(dealer_name)
        html2 = (f'<div style="margin-bottom:20px;">{summary_html}</div>')
        self.dealer_details_label.setText(html2)


    def generate_dealer_progress_table_html(self, dealer_name):
        """
        Generate two separate HTML tables showing progress summary for all personnel in a specific dealer:
        - After-Sales progress table (only after-sales columns)
        - Sales progress table (only sales columns)
        RTL layout with same filtering and exclusion rules.
        """
        dealer_personnel = self.raw[self.raw['عنوان نمایندگی'] == dealer_name]
        
        after_summary = []
        sales_summary = []
        
        after_percentages = []
        sales_percentages = []
        
        seen_combinations = set()
        
        for _, row in dealer_personnel.iterrows():
            name = row['نام و نام خانوادگی']
            main_pos = row.get('عنوان شغل', '')
            alt_pos = row.get('شغل موازی (ارتقا)', '')
            pcode = row.get('کد پرسنلی', '')
            rawcompany = row.get('company', '')
            
            positions_to_add = []
            if pd.notna(main_pos) and main_pos.strip():
                positions_to_add.append(main_pos.strip())
            if pd.notna(alt_pos) and alt_pos.strip():
                alt_positions = [p.strip() for p in alt_pos.split('&&&') if p.strip()]
                positions_to_add.extend(alt_positions)
            if not positions_to_add:
                positions_to_add = ['بدون سمت']
            
            for pos_text in positions_to_add:
                combination_key = (name, pos_text, pcode)
                if combination_key in seen_combinations:
                    continue
                seen_combinations.add(combination_key)
                
                if pos_text not in self.position_mapping:
                    continue
                
                mapped_position = self.position_mapping.get(pos_text, pos_text)
                mapped_company = self.company_mapping.get(rawcompany, rawcompany)
                mapped_categories = [self.car_mapping.get(cat, cat) for cat in self.get_dealer_categories(dealer_name)]
                
                after_requirements = self.get_matching_after_rows(mapped_company, mapped_position, mapped_categories)
                sales_requirements = self.get_matching_sales_rows(mapped_company, mapped_position)
                
                filtered = self.raw[(self.raw['کد پرسنلی'] == pcode) & (self.raw['عنوان نمایندگی'] == dealer_name)]
                passed_courses = filtered['عنوان دوره'].dropna().unique().tolist()
                passed_courses = [self.course_mapping.get(c, c) for c in passed_courses]
                passed_set = set(passed_courses)
                
                grouped = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
                
                for req in after_requirements:
                    criteria = req.get("نام سرفصل", "").strip()
                    car = req.get("نام خودرو", "").strip() or "عمومی"
                    course = req.get("نام دوره آموزشی", "").strip()
                    if criteria and criteria != "nan" and course and course != "nan":
                        grouped["after"][car][criteria].append(course)
                
                sales_df = self.sales_sheets.get(mapped_company)
                if sales_df is not None:
                    for _, r in sales_df.iterrows():
                        row_pos = str(r.get("پست کاری", "")).strip()
                        if row_pos == mapped_position:
                            course = str(r.get("نام دوره آموزشی", "")).strip()
                            criteria = str(r.get("نام سرفصل", "")).strip()
                            car = "فروش"
                            if criteria and criteria != "nan" and course and course != "nan":
                                grouped["sales"][car][criteria].append(course)
                
                pass_status = defaultdict(lambda: defaultdict(dict))
                
                for file_name, cars in grouped.items():
                    for car, criteria_dict in cars.items():
                        for crit, courses in criteria_dict.items():
                            passed = any(c in passed_set for c in courses) or ("گازسوز" in crit)
                            pass_status[file_name][car][crit] = passed
                        for crit in criteria_dict.keys():
                            if "ابزار مخصوص" in crit:
                                others = [c for c in criteria_dict.keys() if c != crit]
                                if all(pass_status[file_name][car].get(c, False) for c in others):
                                    pass_status[file_name][car][crit] = True
                                else:
                                    pass_status[file_name][car][crit] = False
                
                after_total = after_passed = 0
                sales_total = sales_passed = 0
                
                for file_name, cars in grouped.items():
                    for car, crits in cars.items():
                        for crit in crits.keys():
                            if not crit or not crits[crit] or all(not c for c in crits[crit]):
                                continue
                            if file_name == "after":
                                after_total += 1
                                if pass_status[file_name][car].get(crit, False):
                                    after_passed += 1
                            elif file_name == "sales":
                                sales_total += 1
                                if pass_status[file_name][car].get(crit, False):
                                    sales_passed += 1
                
                # Position-based exclusion logic
                show_after = True
                show_sales = True
                
                if "خدمات" in pos_text:
                    show_sales = False
                if "فروش" in pos_text:
                    show_after = False
                
                after_percent = (after_passed / after_total * 100) if after_total > 0 else None
                sales_percent = (sales_passed / sales_total * 100) if sales_total > 0 else None
                
                after_progress = f"{after_percent:.1f}% ({after_passed}/{after_total})" if (after_percent is not None and show_after) else "—"
                sales_progress = f"{sales_percent:.1f}% ({sales_passed}/{sales_total})" if (sales_percent is not None and show_sales) else "—"
                
                if show_after and after_percent is not None:
                    after_percentages.append(after_percent)
                    after_summary.append({
                        'name': name,
                        'position': pos_text,
                        'progress': after_progress,
                    })
                if show_sales and sales_percent is not None:
                    sales_percentages.append(sales_percent)
                    sales_summary.append({
                        'name': name,
                        'position': pos_text,
                        'progress': sales_progress,
                    })
        
        after_summary.sort(key=lambda x: x['name'])
        sales_summary.sort(key=lambda x: x['name'])
        
        avg_after_percent = sum(after_percentages) / len(after_percentages) if after_percentages else 0
        avg_sales_percent = sum(sales_percentages) / len(sales_percentages) if sales_percentages else 0
        
        total_after_progress = f"{avg_after_percent:.1f}% (میانگین)" if after_percentages else "—"
        total_sales_progress = f"{avg_sales_percent:.1f}% (میانگین)" if sales_percentages else "—"
        
        dealer_title = dealer_name[5:]
        
        def build_table(summary, title, col_title):
            parts = [
                f'<div dir="rtl" style="font-family: Vazirmatn, Tahoma, sans-serif; text-align: right; margin: 20px 0;">',
                f'<h2 style="color: #1a237e; font-size: 18px; margin-bottom: 6px;">نمایندگی: {dealer_title}</h2>',
                f'<h3 style="color: #444; font-size: 15px; margin-bottom: 12px;">جدول پیشرفت پرسنل - {title}</h3>',
                '<table style="width: 100%; border-collapse: separate; border-spacing: 0; font-size: 14px; box-shadow: 0 4px 12px rgba(0,0,0,0.06); border-radius: 12px; overflow: hidden;">',
                '<thead>',
                '<tr style="background-color: #1e3a8a; color: #fff;">',
                '<th style="padding: 12px 16px; text-align: right; width: 45%;">نام پرسنل</th>',
                '<th style="padding: 12px 16px; text-align: center; width: 25%;">سمت</th>',
                f'<th style="padding: 12px 16px; text-align: center; width: 30%;">{col_title}</th>',
                '</tr>',
                '</thead>',
                '<tbody>'
            ]
            for i, record in enumerate(summary):
                progress_colored = self._format_progress_with_color(record['progress'])
                row_color = "#ffffff" if i % 2 == 0 else "#f1f5f9"
                parts.extend([
                    f'<tr style="background-color: {row_color}; transition: background-color 0.3s;">',
                    f'<td style="padding: 10px 16px; text-align: right;">{record["name"]}</td>',
                    f'<td style="padding: 10px 16px; text-align: center;">{record["position"]}</td>',
                    f'<td style="padding: 10px 16px; text-align: center;">{progress_colored}</td>',
                    '</tr>'
                ])
            parts.extend([
                '<tr style="background-color: #e0f2fe; font-weight: bold;">',
                '<td style="padding: 10px 16px; text-align: right;">میانگین</td>',
                '<td style="padding: 10px 16px; text-align: center;">همه سمت‌ها</td>',
                f'<td style="padding: 10px 16px; text-align: center;">{total_after_progress if title=="خدمات پس از فروش" else total_sales_progress}</td>',
                '</tr>',
                '</tbody>',
                '</table>',
                '</div>'
            ])
            return ''.join(parts)
        
        after_table_html = build_table(after_summary, "خدمات پس از فروش", "خدمات پس از فروش")
        sales_table_html = build_table(sales_summary, "فروش", "فروش")
        
        return after_table_html, sales_table_html



    def _format_progress_with_color(self, progress_str):
        """Add color coding to progress percentages"""
        if '%' not in progress_str:
            return progress_str
            
        try:
            percent = float(progress_str.split('%')[0].strip())
            if percent >= 90:
                color = "#10b981"  # green
            elif percent >= 70:
                color = "#f59e0b"  # amber
            else:
                color = "#ef4444"  # red
                
            return f'<span style="color:{color}; font-weight:bold;">{progress_str}</span>'
        except:
            return progress_str

def format_worksheet(ws):
    """Apply all requested formatting to an Excel worksheet with working conditional formatting"""
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['D'].width = 20

    # Apply conditional formatting using direct cell styling approach
    max_row = ws.max_row
    if max_row > 1:  # Only apply if there's data beyond header
        for row in range(2, max_row + 1):  # Start from row 2 (skip header)
            cell = ws[f'G{row}']  # Column G is "گذرانده شده"
            if cell.value == 'بله':
                # Light green background
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill(start_color='A8D5A2', end_color='A8D5A2', fill_type='solid')
            elif cell.value == 'خیر':
                # Light red background
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill(start_color='F8A5A5', end_color='F8A5A5', fill_type='solid')
    
    # Set right-to-left direction
    ws.sheet_view.rightToLeft = True